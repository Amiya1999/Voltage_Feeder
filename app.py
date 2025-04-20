# import streamlit as st
# import pandas as pd
# from pyxlsb import open_workbook
# from io import BytesIO
# import plotly.express as px

# st.set_page_config(page_title="⚡ Voltage Analyzer", layout="wide")
# st.title("⚡ Voltage Abnormal Trend Analyzer with Master Data")
# st.markdown("Upload multiple `.xlsb` files + one static **Master Excel** to generate an enhanced summary report.")

# # 🔼 Upload abnormal voltage daily files
# uploaded_files = st.file_uploader("Upload daily `.xlsb` voltage files", type="xlsb", accept_multiple_files=True)

# # 📚 Upload static master sheet (once)
# master_file = st.file_uploader("Upload static Master Data Excel", type=["xlsx", "xls"])

# # ✅ Extract phase-wise voltage abnormality
# def extract_phasewise_uv_ov(file):
#     try:
#         with open_workbook(file) as wb:
#             sheet = wb.get_sheet(wb.sheets[0])
#             data = [row for row in sheet.rows()]
#     except Exception as e:
#         st.error(f"❌ Failed to read {file.name}: {e}")
#         return None

#     rows = [[cell.v for cell in row] for row in data[2:]]
#     records = []

#     for row in rows:
#         try:
#             bus_meter_no = str(row[7]).strip()
#             from_date = pd.to_datetime(row[1], unit='d', origin='1899-12-30')

#             uv_r = float(row[11]) if row[11] is not None else 0.0
#             uv_y = float(row[12]) if row[12] is not None else 0.0
#             uv_b = float(row[13]) if row[13] is not None else 0.0

#             ov_r = float(row[14]) if row[14] is not None else 0.0
#             ov_y = float(row[15]) if row[15] is not None else 0.0
#             ov_b = float(row[16]) if row[16] is not None else 0.0

#             is_uv = any([uv_r > 0, uv_y > 0, uv_b > 0])
#             is_ov = any([ov_r > 0, ov_y > 0, ov_b > 0])
#             is_abnormal = is_uv or is_ov

#             abnormal_phases = []
#             if uv_r > 0 or ov_r > 0: abnormal_phases.append("R")
#             if uv_y > 0 or ov_y > 0: abnormal_phases.append("Y")
#             if uv_b > 0 or ov_b > 0: abnormal_phases.append("B")

#             if is_abnormal:
#                 for phase, uv, ov in zip(['R', 'Y', 'B'], [uv_r, uv_y, uv_b], [ov_r, ov_y, ov_b]):
#                     if uv > 0 or ov > 0:
#                         records.append({
#                             "BUS_METER_NO": bus_meter_no,
#                             "DATE": from_date,
#                             "IS_UNDER": uv > 0,
#                             "IS_OVER": ov > 0,
#                             "PHASE": phase,
#                             "TYPE": "UV" if uv > 0 else "OV",
#                             "VALUE": uv if uv > 0 else ov,
#                             "UV_R": uv_r,
#                             "UV_Y": uv_y,
#                             "UV_B": uv_b,
#                             "OV_R": ov_r,
#                             "OV_Y": ov_y,
#                             "OV_B": ov_b,
#                             "PHASES": ",".join(abnormal_phases)
#                         })
#         except:
#             continue

#     return pd.DataFrame(records)

# if uploaded_files and master_file:
#     # ✅ Step 1: Load master sheet
#     master_df = pd.read_excel(master_file, sheet_name=0)
#     master_df.columns = master_df.columns.str.strip()
#     master_df["Connected Energy Meter"] = master_df["Connected Energy Meter"].astype(str).str.strip()

#     # ✅ Step 2: Process uploaded voltage files
#     all_abnormal_data = []

#     for file in uploaded_files:
#         df = extract_phasewise_uv_ov(file)
#         if df is not None and not df.empty:
#             all_abnormal_data.append(df)
#         else:
#             st.warning(f"⚠️ Skipping file: {file.name} (no abnormal rows)")

#     if all_abnormal_data:
#         full_df = pd.concat(all_abnormal_data, ignore_index=True)

#         # Step 3: Create summarized trends
#         summary = full_df.groupby("BUS_METER_NO").agg(
#             Days_Abnormal=('DATE', 'nunique'),
#             From_Date=('DATE', 'min'),
#             To_Date=('DATE', 'max'),
#             Days_Under=('IS_UNDER', 'sum'),
#             Days_Over=('IS_OVER', 'sum'),
#             Phases_Affected=('PHASES', lambda x: ', '.join(sorted(set(','.join(x).split(','))))),
#             UV_R_Avg=('UV_R', 'mean'),
#             UV_Y_Avg=('UV_Y', 'mean'),
#             UV_B_Avg=('UV_B', 'mean'),
#             OV_R_Avg=('OV_R', 'mean'),
#             OV_Y_Avg=('OV_Y', 'mean'),
#             OV_B_Avg=('OV_B', 'mean')
#         ).reset_index()

#         # Convert averages to percentage
#         percent_cols = ['UV_R_Avg', 'UV_Y_Avg', 'UV_B_Avg', 'OV_R_Avg', 'OV_Y_Avg', 'OV_B_Avg']
#         for col in percent_cols:
#             summary[col] = (summary[col] * 100).round(2)

#         # ✅ Step 4: Merge with master data
#         merged_summary = summary.merge(
#             master_df, how="left",
#             left_on="BUS_METER_NO",
#             right_on="Connected Energy Meter"
#         )

#         st.success("✅ Final master-linked voltage trend summary generated!")
#         st.dataframe(merged_summary, use_container_width=True)

#         # ✅ Step 5: Download final Excel
#         output = BytesIO()
#         with pd.ExcelWriter(output, engine='openpyxl') as writer:
#             merged_summary.to_excel(writer, index=False, sheet_name="Final_Summary")
#             full_df.to_excel(writer, index=False, sheet_name="Raw_Abnormal_Log")
#         output.seek(0)

#         st.download_button(
#             label="📥 Download Final Excel Report",
#             data=output,
#             file_name="Voltage_Trend_Summary_With_Master.xlsx",
#             mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#         )

#         # ✅ Step 6: Interactive Chart with Percentage Conversion
#         st.markdown("### 📊 View Phase-wise Graph")
#         meter_list = merged_summary["BUS_METER_NO"].dropna().unique().tolist()
#         selected_meter = st.selectbox("Select a Meter", meter_list)

#         meter_data = full_df[full_df["BUS_METER_NO"] == selected_meter]
#         if not meter_data.empty:
#             meter_data = meter_data.sort_values("DATE")
#             meter_data["VALUE_PERCENT"] = (meter_data["VALUE"] * 100).round(2)
#             color_map = {"UV": "red", "OV": "blue"}

#             fig = px.bar(
#                 meter_data,
#                 x="DATE",
#                 y="VALUE_PERCENT",
#                 color="TYPE",
#                 facet_col="PHASE",
#                 color_discrete_map=color_map,
#                 title=f"Voltage Deviation by Phase — {selected_meter}",
#                 labels={"VALUE_PERCENT": "Voltage (%)"},
#                 height=450
#             )
#             st.plotly_chart(fig, use_container_width=True)

#     else:
#         st.warning("⚠️ No abnormal readings found in uploaded files.")


import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

st.set_page_config(page_title="⚡ Voltage Analyzer", layout="wide")
st.title("⚡ Voltage Abnormal Trend Analyzer with Chart-Embedded Excel")
st.markdown("Upload daily `.xlsb` or `.xlsx` voltage files and one static **Master Excel** to generate a full summary report with per-meter charts.")

# Accept both xlsb and xlsx files
uploaded_files = st.file_uploader(
    "Upload daily `.xlsb` or `.xlsx` voltage files",
    type=["xlsb", "xlsx"],
    accept_multiple_files=True
)

master_file = st.file_uploader("Upload static Master Data Excel", type=["xlsx", "xls"])

# ✅ Extract phase-wise voltage abnormality from either format
def extract_phasewise_uv_ov(file):
    try:
        if file.name.endswith(".xlsb"):
            from pyxlsb import open_workbook
            with open_workbook(file) as wb:
                sheet = wb.get_sheet(wb.sheets[0])
                data = [row for row in sheet.rows()]
            rows = [[cell.v for cell in row] for row in data[2:]]

        elif file.name.endswith(".xlsx"):
            df = pd.read_excel(file, sheet_name=0, header=None, skiprows=2)
            rows = df.values.tolist()
        else:
            st.warning(f"❌ Unsupported file format: {file.name}")
            return None

        records = []
        for row in rows:
            try:
                bus_meter_no = str(row[7]).strip()
                from_date = pd.to_datetime(row[1], unit='d', origin='1899-12-30')

                uv_r = float(row[11]) if row[11] is not None else 0.0
                uv_y = float(row[12]) if row[12] is not None else 0.0
                uv_b = float(row[13]) if row[13] is not None else 0.0

                ov_r = float(row[14]) if row[14] is not None else 0.0
                ov_y = float(row[15]) if row[15] is not None else 0.0
                ov_b = float(row[16]) if row[16] is not None else 0.0

                if any([uv_r, uv_y, uv_b, ov_r, ov_y, ov_b]):
                    records.append({
                        "BUS_METER_NO": bus_meter_no,
                        "DATE": from_date,
                        "UV_R": uv_r,
                        "UV_Y": uv_y,
                        "UV_B": uv_b,
                        "OV_R": ov_r,
                        "OV_Y": ov_y,
                        "OV_B": ov_b
                    })
            except:
                continue

        return pd.DataFrame(records)

    except Exception as e:
        st.error(f"⚠️ Failed to read {file.name}: {e}")
        return None

# 🧠 Processing
if uploaded_files and master_file:
    master_df = pd.read_excel(master_file)
    master_df.columns = master_df.columns.str.strip()
    master_df["Connected Energy Meter"] = master_df["Connected Energy Meter"].astype(str).str.strip()

    all_abnormal_data = []
    for file in uploaded_files:
        df = extract_phasewise_uv_ov(file)
        if df is not None and not df.empty:
            all_abnormal_data.append(df)
        else:
            st.warning(f"⚠️ Skipping file: {file.name} (no abnormal rows)")

    if all_abnormal_data:
        full_df = pd.concat(all_abnormal_data, ignore_index=True)

        # Add boolean flags
        full_df["IS_UNDER"] = full_df[["UV_R", "UV_Y", "UV_B"]].gt(0).any(axis=1)
        full_df["IS_OVER"] = full_df[["OV_R", "OV_Y", "OV_B"]].gt(0).any(axis=1)

        # Summary
        summary = full_df.groupby("BUS_METER_NO").agg(
            Days_Abnormal=('DATE', 'nunique'),
            From_Date=('DATE', 'min'),
            To_Date=('DATE', 'max'),
            Days_Under=('IS_UNDER', lambda x: full_df.loc[x.index].groupby("DATE")["IS_UNDER"].any().sum()),
            Days_Over=('IS_OVER', lambda x: full_df.loc[x.index].groupby("DATE")["IS_OVER"].any().sum()),
            UV_R_Avg=('UV_R', 'mean'),
            UV_Y_Avg=('UV_Y', 'mean'),
            UV_B_Avg=('UV_B', 'mean'),
            OV_R_Avg=('OV_R', 'mean'),
            OV_Y_Avg=('OV_Y', 'mean'),
            OV_B_Avg=('OV_B', 'mean')
        ).reset_index()

        # Percent conversion
        percent_cols = ['UV_R_Avg', 'UV_Y_Avg', 'UV_B_Avg', 'OV_R_Avg', 'OV_Y_Avg', 'OV_B_Avg']
        for col in percent_cols:
            summary[col] = (summary[col] * 100).round(2)

        # Merge with master
        merged_summary = summary.merge(
            master_df, how="left",
            left_on="BUS_METER_NO",
            right_on="Connected Energy Meter"
        )

        # Create Excel
        output = BytesIO()
        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = "Final_Summary"
        for r in dataframe_to_rows(merged_summary, index=False, header=True):
            ws_summary.append(r)

        # Sheet 2: Raw Log
        ws_raw = wb.create_sheet("Raw_Abnormal_Log")
        for r in dataframe_to_rows(full_df, index=False, header=True):
            ws_raw.append(r)

        # Sheet 3: Charts
        ws_chart = wb.create_sheet("Charts")
        row_cursor = 1
        for meter in full_df["BUS_METER_NO"].unique():
            meter_data = full_df[full_df["BUS_METER_NO"] == meter].sort_values("DATE")
            meter_data_chart = meter_data[["DATE", "UV_R", "UV_Y", "UV_B", "OV_R", "OV_Y", "OV_B"]].copy()
            meter_data_chart[["UV_R", "UV_Y", "UV_B", "OV_R", "OV_Y", "OV_B"]] *= 100

            def classify_type(row):
                is_uv = any([row["UV_R"] > 0, row["UV_Y"] > 0, row["UV_B"] > 0])
                is_ov = any([row["OV_R"] > 0, row["OV_Y"] > 0, row["OV_B"] > 0])
                if is_uv and is_ov:
                    return "Both"
                elif is_uv:
                    return "Under Voltage"
                elif is_ov:
                    return "Over Voltage"
                return "Normal"

            meter_data_chart["TYPE"] = meter_data_chart.apply(classify_type, axis=1)
            meter_data_chart = meter_data_chart[["DATE", "TYPE", "UV_R", "UV_Y", "UV_B", "OV_R", "OV_Y", "OV_B"]]

            ws_chart.cell(row=row_cursor, column=1, value=f"BUS_METER_NO: {meter}")
            row_cursor += 1

            for r in dataframe_to_rows(meter_data_chart, index=False, header=True):
                ws_chart.append(r)

            data_rows = len(meter_data_chart) + 1
            chart_start_row = row_cursor
            chart_end_row = chart_start_row + data_rows - 1

            chart = BarChart()
            chart.title = f"{meter} - Voltage Chart (%)"
            chart.y_axis.title = "Voltage (%)"
            chart.x_axis.title = "Date"
            data_ref = Reference(ws_chart, min_col=3, max_col=8, min_row=chart_start_row, max_row=chart_end_row)
            cats_ref = Reference(ws_chart, min_col=1, min_row=chart_start_row + 1, max_row=chart_end_row)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)

            ws_chart.add_chart(chart, f"J{chart_start_row}")
            row_cursor = chart_end_row + 20

        wb.save(output)
        output.seek(0)

        st.success("✅ Excel report with .xlsb/.xlsx support, proper summary and clean charts is ready!")
        st.download_button(
            label="📥 Download Voltage Summary with Charts",
            data=output,
            file_name="Voltage_Analyzer_With_Charts.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.warning("⚠️ No abnormal data found in the uploaded files.")
